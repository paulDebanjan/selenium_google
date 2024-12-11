from datetime import datetime as timesite
from  openpyxl import load_workbook,Workbook
import web_operation


excel_file_path = 'Excel.xlsx'
iteration_count = 0

# to get day name from user given date
def get_day_name(number_formate_date):   
    date_object = timesite.strptime(number_formate_date, '%d%m%Y')
    day_name = date_object.strftime("%A")
    return day_name

# operate all excel opeation from this function
def excel_operation(day):
    data_collected_column = 'C'                         # excel column name
    first_col = 3                                       # excel column data featching start porsition
    row_number = first_col
    workbook = load_workbook(excel_file_path)           # loading excel file
    selected_sheet = workbook[day]                      # selecting worksheet depend on day name
    last_column = len(selected_sheet[data_collected_column])            # masurement column length for getting all data
    rng = selected_sheet[f"{data_collected_column}{first_col}:{data_collected_column}{last_column}"]        # featching data
    result_sheet_path =crate_resule_excel_file(day)
    for cells in rng:                                                                                   # for print data
        for cell in cells:
            longest_option, shortest_option = web_operation.web_to_scripting_operation(cell.value)
            print(f"Longest path: {longest_option}")
            print(f"shortest_option : {shortest_option}")
        excel_data_save(result_sheet_path, day,row_number,longest_option,shortest_option)
        row_number += 1

def crate_resule_excel_file(day):
    result_sheet_path = 'result/result_sheet.xlsx'
    source_workbook = load_workbook('Excel.xlsx')
    source_sheet = source_workbook[day]
    
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = source_sheet.title
    

    for row in source_sheet.iter_rows(values_only = True):
        new_sheet.append(row)

    new_workbook.save(result_sheet_path)
    return result_sheet_path

def excel_data_save(result_sheet_path,day,row_num,longest_option,shortest_option):
    longest_option_col = 'D'
    shortest_option_col = 'E'
    print(f"row Number: {row_num}")
    workbook = load_workbook(result_sheet_path)           # loading excel file
    selected_sheet = workbook[day]
    selected_sheet[f"{longest_option_col}{row_num}"] = longest_option
    selected_sheet[f"{shortest_option_col}{row_num}"] = shortest_option
    workbook.save(result_sheet_path)
# main operation function
def main():
    user_input_date = input('Enter a Date (DDMMYYYY): ')
    day = get_day_name(user_input_date)                     #sending user giving data for geting week day
    print(f"\nYour Selected Day is: '{day}' \nand this day selected values are: \n")
    excel_operation(day)

main()